'''
Author: Declan Hollingworth
Date: 2020-06-14
This function searches for a specific key word throughout an excel column and replaces the text with a new given string.
Originally made for a work project and simplified to a general function for GitHub.
'''
from openpyxl import load_workbook

path = 'D:\Python Excel\test_sheet.xlsx'
wb = load_workbook(path)
ws = wb.active

# Text To Find
prev_text = "Text To Find"

# Text to replace with
new_text = "New Text"

# Title of new xlsx document
save_title = "New Excel Title"


def double_text_change(prev_text, new_text, save_title):
    column1 = "A"
    i = 1
    for _ in range (ws.max_row):
        cellA = str(column1) + str(i)
        #print(cellA)
        if ws[cellA].value == prev_text:
            ws[cellA] = new_text
        i += 1
    title = save_title + '.xlsx'
    wb.save(title)
    print("Finished!")

double_text_change(prev_text, new_text, save_title)